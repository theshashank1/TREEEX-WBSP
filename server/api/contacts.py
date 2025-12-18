"""
Contact Management API endpoints for WhatsApp Business.
Supports CRUD operations and CSV/Excel import for contacts with labels.
"""

import csv
import io
import re
from datetime import datetime, timezone
from typing import Annotated, List, Optional
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from server.core.db import get_async_session
from server.core.monitoring import log_event, log_exception
from server.dependencies import User, get_current_user, get_workspace_member
from server.models.contacts import Contact

router = APIRouter(prefix="/contacts", tags=["Contacts"])

# Type aliases for dependencies
SessionDep = Annotated[AsyncSession, Depends(get_async_session)]
CurrentUserDep = Annotated[User, Depends(get_current_user)]


# ============================================================================
# CONSTANTS
# ============================================================================

# E.164 phone number regex
E164_REGEX = re.compile(r"^\+[1-9]\d{1,14}$")

# Maximum contacts per import
MAX_IMPORT_ROWS = 10000


# ============================================================================
# SCHEMAS
# ============================================================================


class ContactCreate(BaseModel):
    """Schema for creating a new contact"""

    workspace_id: UUID
    phone_number: str = Field(..., description="Phone number in E.164 format (e.g., +15551234567)")
    name: Optional[str] = Field(None, max_length=255)
    tags: Optional[List[str]] = Field(default_factory=list, description="Labels/tags for the contact")

    @field_validator("phone_number")
    @classmethod
    def validate_phone_number(cls, v: str) -> str:
        v = v.strip()
        if not E164_REGEX.match(v):
            raise ValueError(
                "Phone number must be in E.164 format (e.g., +15551234567)"
            )
        return v


class ContactUpdate(BaseModel):
    """Schema for updating a contact"""

    name: Optional[str] = Field(None, max_length=255)
    tags: Optional[List[str]] = None
    opted_in: Optional[bool] = None


class ContactResponse(BaseModel):
    """Schema for contact response"""

    id: UUID
    workspace_id: UUID
    wa_id: str
    phone_number: str
    name: Optional[str]
    opted_in: bool
    tags: Optional[List[str]]
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


class ContactListResponse(BaseModel):
    """Schema for paginated contact list"""

    data: List[ContactResponse]
    total: int
    limit: int
    offset: int


class ImportRowResult(BaseModel):
    """Result for a single import row"""

    row_number: int
    phone_number: Optional[str]
    status: str  # "imported", "updated", "failed"
    reason: Optional[str] = None


class ImportResponse(BaseModel):
    """Schema for import response"""

    total_rows: int
    imported: int
    updated: int
    failed: int
    results: List[ImportRowResult]


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================


def normalize_phone_number(phone: str) -> str:
    """
    Normalize phone number to E.164 format.
    Removes spaces, dashes, parentheses.
    """
    # Remove common formatting characters
    cleaned = re.sub(r"[\s\-\(\)\.]", "", phone)

    # Ensure it starts with +
    if not cleaned.startswith("+"):
        cleaned = "+" + cleaned

    return cleaned


def parse_labels(labels_str: Optional[str]) -> List[str]:
    """Parse comma or semicolon separated labels string."""
    if not labels_str:
        return []

    # Split by comma or semicolon
    labels = re.split(r"[,;]", labels_str)

    # Clean up each label
    return [label.strip() for label in labels if label.strip()]


# ============================================================================
# ENDPOINTS
# ============================================================================


@router.post("", response_model=ContactResponse, status_code=201)
async def create_contact(
    data: ContactCreate,
    session: SessionDep,
    current_user: CurrentUserDep,
):
    """
    Create a new contact.

    Phone number must be in E.164 format (e.g., +15551234567).
    Requires workspace membership.
    """
    # Verify workspace membership
    await get_workspace_member(data.workspace_id, current_user, session)

    # Check for existing contact with same phone number
    wa_id = data.phone_number.replace("+", "")
    result = await session.execute(
        select(Contact).where(
            and_(
                Contact.workspace_id == data.workspace_id,
                Contact.wa_id == wa_id,
                Contact.deleted_at.is_(None),
            )
        )
    )
    existing = result.scalar_one_or_none()

    if existing:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "CONTACT_EXISTS",
                "message": "A contact with this phone number already exists",
            },
        )

    # Create contact
    contact = Contact(
        workspace_id=data.workspace_id,
        wa_id=wa_id,
        phone_number=data.phone_number,
        name=data.name,
        tags=data.tags or [],
        opted_in=True,
        opt_in_source="api",
        opt_in_date=datetime.now(timezone.utc).replace(tzinfo=None),
    )

    session.add(contact)
    await session.commit()
    await session.refresh(contact)

    log_event(
        "contact_created",
        contact_id=str(contact.id),
        workspace_id=str(data.workspace_id),
    )

    return contact


@router.get("", response_model=ContactListResponse)
async def list_contacts(
    session: SessionDep,
    current_user: CurrentUserDep,
    workspace_id: UUID = Query(..., description="Workspace ID"),
    tags: Optional[str] = Query(None, description="Filter by tags (comma-separated)"),
    opted_in: Optional[bool] = Query(None, description="Filter by opt-in status"),
    search: Optional[str] = Query(None, description="Search by name or phone"),
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
):
    """
    List contacts for a workspace.

    Supports filtering by tags, opt-in status, and search.
    Requires workspace membership.
    """
    # Verify workspace membership
    await get_workspace_member(workspace_id, current_user, session)

    # Build query
    query = select(Contact).where(
        Contact.workspace_id == workspace_id,
        Contact.deleted_at.is_(None),
    )

    # Filter by tags
    if tags:
        tag_list = [t.strip() for t in tags.split(",") if t.strip()]
        if tag_list:
            # Use overlap operator for array intersection
            query = query.where(Contact.tags.overlap(tag_list))

    # Filter by opt-in status
    if opted_in is not None:
        query = query.where(Contact.opted_in == opted_in)

    # Search filter
    if search:
        search_term = f"%{search}%"
        query = query.where(
            or_(
                Contact.name.ilike(search_term),
                Contact.phone_number.ilike(search_term),
            )
        )

    # Get total count
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await session.execute(count_query)
    total = total_result.scalar() or 0

    # Apply pagination
    query = query.order_by(Contact.created_at.desc()).offset(offset).limit(limit)
    result = await session.execute(query)
    contacts = result.scalars().all()

    return ContactListResponse(
        data=[ContactResponse.model_validate(c) for c in contacts],
        total=total,
        limit=limit,
        offset=offset,
    )


@router.get("/{contact_id}", response_model=ContactResponse)
async def get_contact(
    contact_id: UUID,
    session: SessionDep,
    current_user: CurrentUserDep,
):
    """
    Get contact details.

    Requires workspace membership.
    """
    result = await session.execute(
        select(Contact).where(
            Contact.id == contact_id,
            Contact.deleted_at.is_(None),
        )
    )
    contact = result.scalar_one_or_none()

    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")

    # Verify workspace membership
    await get_workspace_member(contact.workspace_id, current_user, session)

    return contact


@router.patch("/{contact_id}", response_model=ContactResponse)
async def update_contact(
    contact_id: UUID,
    data: ContactUpdate,
    session: SessionDep,
    current_user: CurrentUserDep,
):
    """
    Update a contact.

    Requires workspace membership.
    """
    result = await session.execute(
        select(Contact).where(
            Contact.id == contact_id,
            Contact.deleted_at.is_(None),
        )
    )
    contact = result.scalar_one_or_none()

    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")

    # Verify workspace membership
    await get_workspace_member(contact.workspace_id, current_user, session)

    # Update fields
    if data.name is not None:
        contact.name = data.name

    if data.tags is not None:
        contact.tags = data.tags

    if data.opted_in is not None:
        contact.opted_in = data.opted_in

    await session.commit()
    await session.refresh(contact)

    log_event(
        "contact_updated",
        contact_id=str(contact_id),
    )

    return contact


@router.delete("/{contact_id}", status_code=204)
async def delete_contact(
    contact_id: UUID,
    session: SessionDep,
    current_user: CurrentUserDep,
):
    """
    Soft delete a contact.

    Requires workspace membership.
    """
    result = await session.execute(
        select(Contact).where(
            Contact.id == contact_id,
            Contact.deleted_at.is_(None),
        )
    )
    contact = result.scalar_one_or_none()

    if not contact:
        raise HTTPException(status_code=404, detail="Contact not found")

    # Verify workspace membership
    await get_workspace_member(contact.workspace_id, current_user, session)

    # Soft delete
    contact.soft_delete()
    await session.commit()

    log_event(
        "contact_deleted",
        contact_id=str(contact_id),
    )

    return None


@router.post("/import", response_model=ImportResponse)
async def import_contacts(
    session: SessionDep,
    current_user: CurrentUserDep,
    workspace_id: UUID = Query(..., description="Workspace ID"),
    file: UploadFile = File(..., description="CSV or Excel file"),
):
    """
    Import contacts from CSV or Excel file.

    Expected columns:
    - phone (required): Phone number in E.164 format or common formats
    - name (optional): Contact name
    - labels/tags (optional): Comma or semicolon separated labels

    Returns per-row import status.
    Requires workspace membership.
    """
    # Verify workspace membership
    await get_workspace_member(workspace_id, current_user, session)

    # Read file content
    try:
        content = await file.read()
    except Exception as e:
        log_exception("import_read_error", e)
        raise HTTPException(
            status_code=400,
            detail={"code": "FILE_READ_ERROR", "message": "Failed to read file"},
        )

    # Detect file type
    filename = file.filename or ""
    is_excel = filename.lower().endswith((".xlsx", ".xls"))
    is_csv = filename.lower().endswith(".csv") or not is_excel

    rows = []

    if is_excel:
        # Parse Excel
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
            ws = wb.active

            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value).lower().strip() if cell.value else "")

            # Parse data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_dict[headers[i]] = str(value) if value is not None else ""
                if any(row_dict.values()):  # Skip empty rows
                    rows.append((row_idx, row_dict))

        except ImportError:
            raise HTTPException(
                status_code=400,
                detail={
                    "code": "EXCEL_NOT_SUPPORTED",
                    "message": "Excel parsing requires openpyxl package. Please use CSV format.",
                },
            )
        except Exception as e:
            log_exception("import_excel_parse_error", e)
            raise HTTPException(
                status_code=400,
                detail={"code": "PARSE_ERROR", "message": "Failed to parse Excel file"},
            )
    else:
        # Parse CSV
        try:
            text = content.decode("utf-8-sig")  # Handle BOM
            reader = csv.DictReader(io.StringIO(text))

            # Normalize headers
            if reader.fieldnames:
                normalized_fieldnames = [h.lower().strip() for h in reader.fieldnames]
                reader.fieldnames = normalized_fieldnames

            for row_idx, row in enumerate(reader, start=2):
                rows.append((row_idx, row))

        except Exception as e:
            log_exception("import_csv_parse_error", e)
            raise HTTPException(
                status_code=400,
                detail={"code": "PARSE_ERROR", "message": "Failed to parse CSV file"},
            )

    # Limit rows
    if len(rows) > MAX_IMPORT_ROWS:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "TOO_MANY_ROWS",
                "message": f"Maximum {MAX_IMPORT_ROWS} rows allowed per import",
            },
        )

    # Process rows
    results = []
    imported = 0
    updated = 0
    failed = 0

    for row_num, row in rows:
        # Find phone column
        phone = None
        for key in ["phone", "phone_number", "phonenumber", "mobile", "tel", "telephone"]:
            if key in row and row[key]:
                phone = row[key].strip()
                break

        if not phone:
            results.append(
                ImportRowResult(
                    row_number=row_num,
                    phone_number=None,
                    status="failed",
                    reason="Missing phone number",
                )
            )
            failed += 1
            continue

        # Normalize phone number
        phone = normalize_phone_number(phone)

        # Validate E.164
        if not E164_REGEX.match(phone):
            results.append(
                ImportRowResult(
                    row_number=row_num,
                    phone_number=phone,
                    status="failed",
                    reason="Invalid phone number format (must be E.164)",
                )
            )
            failed += 1
            continue

        # Get name
        name = None
        for key in ["name", "contact_name", "full_name", "fullname"]:
            if key in row and row[key]:
                name = row[key].strip()[:255]
                break

        # Get labels/tags
        labels = []
        for key in ["labels", "tags", "label", "tag", "groups", "group"]:
            if key in row and row[key]:
                labels = parse_labels(row[key])
                break

        # Check for existing contact
        wa_id = phone.replace("+", "")
        result = await session.execute(
            select(Contact).where(
                and_(
                    Contact.workspace_id == workspace_id,
                    Contact.wa_id == wa_id,
                )
            )
        )
        existing = result.scalar_one_or_none()

        try:
            if existing:
                # Update existing contact
                if name:
                    existing.name = name
                if labels:
                    # Merge labels
                    existing_tags = existing.tags or []
                    merged_tags = list(set(existing_tags + labels))
                    existing.tags = merged_tags

                # Restore if soft-deleted
                if existing.deleted_at:
                    existing.deleted_at = None

                results.append(
                    ImportRowResult(
                        row_number=row_num,
                        phone_number=phone,
                        status="updated",
                    )
                )
                updated += 1
            else:
                # Create new contact
                contact = Contact(
                    workspace_id=workspace_id,
                    wa_id=wa_id,
                    phone_number=phone,
                    name=name,
                    tags=labels,
                    opted_in=True,
                    opt_in_source="import",
                    opt_in_date=datetime.now(timezone.utc).replace(tzinfo=None),
                )
                session.add(contact)

                results.append(
                    ImportRowResult(
                        row_number=row_num,
                        phone_number=phone,
                        status="imported",
                    )
                )
                imported += 1

        except Exception as e:
            log_exception("import_row_error", e, row_num=row_num)
            results.append(
                ImportRowResult(
                    row_number=row_num,
                    phone_number=phone,
                    status="failed",
                    reason=str(e),
                )
            )
            failed += 1

    # Commit all changes
    await session.commit()

    log_event(
        "contacts_imported",
        workspace_id=str(workspace_id),
        total=len(rows),
        imported=imported,
        updated=updated,
        failed=failed,
    )

    return ImportResponse(
        total_rows=len(rows),
        imported=imported,
        updated=updated,
        failed=failed,
        results=results,
    )
